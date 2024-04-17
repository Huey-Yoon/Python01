import openpyxl as px      # 엑셀 출력
import os
from datetime import datetime
import re

# 현재 실행 파일 경로 가져오고, 입력파일 지정하기
program_path = os.path.abspath(__file__)
path = os.path.dirname(program_path)
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

# 엑셀 파일 로드
# load_workbook : 지정한 엑셀 파일 읽어와서 WorkBook 객체로 가져온다.
workbook = px.load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']

# 엑셀 출력 객체 생성
output_workbook = px.Workbook()
output_worksheet = output_workbook.active
output_workbook.title = 'out_january_2013'

# 정규표현식으로 필터링
# * Customer Name 이 대문자 J로 시작하는 행만 필터링
# pattern = re.compile(r'(?P<my_pattern>^J.*)')  # 1번 문제.
pattern = re.compile(r'(?P<my_pattern>^R.*)')

index = 1

workbook = px.load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']

data = []
header = [cell.value for cell in worksheet[1]]
data.append(header)

for row in worksheet.iter_rows(min_row=2, values_only=True):
    row_list = []
    # 정규표현식 패턴에 일치하는지 확인
    # row[index] : Customer Name
    if pattern.match( row[index] ):
        for cell_value in row:
            row_list.append(cell_value)
    if row_list:
        data.append(row_list)

# 엑셀 파일 출력
for row_index, row in enumerate(data, 1):
    # 열 반복
    for column_index, value in enumerate(row, 1):
        if type(value) == datetime:
            value = value.strftime('%Y/%m/%d')
            output_worksheet.cell(row=row_index, column=column_index, value=value)
        else:
        # output_worksheet.cell(row=행, column=열, value=값) : 셀의 값을 지정
            output_worksheet.cell(row=row_index, column=column_index, value=value)

# 엑셀 통합 문서 저장
output_workbook.save(output_file)