
from openpyxl import load_workbook      # 엑셀 입력
from openpyxl import Workbook           # 엑셀 출력      

import glob
import os

import tkinter as tk
from tkinter import filedialog

# 입력 파일 선택 버튼 클릭 시,
def open_input_file():
    print('입력 파일 선택...')
    # 파일 선택 상자에서 폴더 경로 가져오기
    directory_name = filedialog.askdirectory()
    if directory_name:
        # 폴더 경로 지정
        input_file.set(directory_name)


# 출력 파일 선택 버튼 클릭 시,
def open_output_file():
    print('출력 파일 선택...')
    # 파일 선택 상자에서 파일이름 가져오기
    filename = filedialog.asksaveasfilename()
    if filename:
        # 파일명 지정
        output_file.set(filename)

# 실행 버튼
def run():
    print('입력 파일 경로 : ' + input_file.get())
    print('출력 파일 경로 : ' + output_file.get())
    print('데이터 분석을 시작합니다...')
    work()

# 데이터 분석
def work():

    # 엑셀 출력 객체 생성
    output_workbook = Workbook()
    output_worksheet =  output_workbook.active  # 워크시트 활성화
    output_worksheet.title = '통합문서 합계평균' # 워크시트 이름 지정

    input_directory = input_file.get()          # 엑셀 파일들이 있는 경로 
    first_worksheet = True

    header = ['엑셀파일', '워크시트', '시트 합계', '시트 평균', '파일 합계', '파일 평균']
    output_worksheet.append(header)

    # 모든 엑셀파일 반복
    for excel_file in glob.glob( os.path.join(input_directory, '*.xlsx') ):
        workbook = load_workbook(filename=excel_file, read_only=True)
        workbook_name = os.path.basename(excel_file)                        

        book_list = []
        list_sum = []           # [시트1합계,시트2합계,시트3합계]
        list_count = []         # [시트1개수,시트2개수,시트3개수]
        # 엑셀 파일의 워크시트 반복
        for worksheet in workbook:
            sheet_list = []
            sheet_list.append(workbook_name)    # 엑셀파일명 추가
            sheet_list.append(worksheet.title)  # 워크시트명 추가

            sheet_sum = 0
            sheet_avg = 0.0
            sheet_row = 0
            # 데이터들을 가져오기
            for row in worksheet.iter_rows(min_row=2):
                # row[3] : Sale Amount
                sale_amount = float( row[3].value )
                sheet_sum += sale_amount        # 합계 연산
                sheet_row += 1                  # 데이터 행 개수 
            sheet_avg = sheet_sum / sheet_row   # 평균 연산
            sheet_list.append(sheet_sum)
            sheet_list.append(sheet_avg)

            list_sum.append(sheet_sum)
            list_count.append(sheet_row)

            book_list.append(sheet_list)
        

        file_sum = sum( list_sum )
        file_avg = file_sum / sum(list_count)

        for item_list in book_list:
            # item_list : [파일명, 시트명, 시트합계, 시트평균]
            item_list.append(file_sum)
            item_list.append(file_avg)
            # 출력 시트에 행 추가
            output_worksheet.append(item_list)

    workbook.close()   


    # 엑셀 통합 문서 저장
    output_workbook.save(output_file.get())




# 윈도우 화면 생성
window = tk.Tk()
window.title('GUI 프로그램')

# 창 크기 지정
window.geometry('600x300')

# 라벨 생성
label = tk.Label(window, text="GUI 프로그램 라벨", padx=20, pady=10)
label.pack()

# 입력 상자 생성 - 입력 파일 경로
input_file = tk.StringVar()         # 문자열 변수 
input_entry = tk.Entry(window, textvariable=input_file, width=100)
input_entry.pack()

# 버튼 - 입력 파일 선택 버튼
input_button = tk.Button(window, text="입력 파일 선택", command=open_input_file)
input_button.pack()


# 입력 상자 생성 - 출력 파일 경로
output_file = tk.StringVar()         # 문자열 변수 
output_entry = tk.Entry(window, textvariable=output_file, width=100)
output_entry.pack()

# 버튼 - 입력 파일 선택 버튼
output_button = tk.Button(window, text="출력 파일 선택", command=open_output_file)
output_button.pack()

# 실행 버튼
run_button = tk.Button(window, text="실행", padx=10, pady=10, command=run)
run_button.pack()

window.mainloop()