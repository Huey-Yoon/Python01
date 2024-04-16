from openpyxl import load_workbook  # 엑셀 입력
from openpyxl import Workbook       # 엑셀 출력
import os

# 현재 실행 파일 경로 가져오고, 입력파일 지정하기
program_path = os.path.abspath(__file__)
path = os.path.dirname(program_path)
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

# 엑셀 파일 로드
# load_workbook : 지정한 엑셀 파일 읽어와서 WorkBook 객체로 가져온다.
workbook = load_workbook(input_file)

# january_2013 워크시트만 입력
worksheet = workbook['january_2013']

# 엑셀 출력 객체 생성
output_workbook = Workbook()
output_worksheet = output_workbook.active   # 워크시트 활성화
output_worksheet.title = 'out_january_2013' # 워크시트 활성화 지정

# sales_2013.xlsx 의 january_2013 워크시트를 반복하여
# output02.xlsx 의 out_january_2013 워크시트로 출력
# 행 반복
for row_index, row in enumerate(worksheet.iter_rows(), 1):
    # 열 반복
    for column_index, cell in enumerate(row, 1):
        # 셀의 날짜 형식을 지정하여 출력하기
        # cell.data_type == 'n'              : 숫자 타입 이면서,
        # cell.number_format == 'mm/dd/yyyy' : 날짜형식 포맷이면,
        # 우리가 원하는 포맷으로 출력하겠다.
        if cell.data_type == 'd' and cell.number_format == 'mm-dd-yy':
            print('날짜형식 변환 : 월-일-연  --->  연/월/일')
            date_cell = cell.value.strftime('%Y/%m/%d') # 연/월/일 형식으로 변환
            output_worksheet.cell(row=row_index, column=column_index, value=date_cell)
        else:
            output_worksheet.cell(row=row_index, column=column_index, value=cell.value)

# 엑셀 통합 문서 지정
output_workbook.save(output_file)

